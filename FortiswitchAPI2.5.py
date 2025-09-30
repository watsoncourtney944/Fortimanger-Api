import requests
import json
import openpyxl
import sys
import glob

file_path2 = "Fortiswitch.xlsx"
def read_excel_data(file_path2):
    """Reads FortiManager details and FortiGate hostnames from an Excel file."""
    workbook = openpyxl.load_workbook(file_path2)
    sheet = workbook.active  

    # Read values from Excel (Assumes first row is headers)
    fortimanager_ip = sheet.cell(row=2, column=2).value  # Column B (FortiManager IP)
    api_token = sheet.cell(row=2, column=3).value  # Column C (API Token)
    adom = sheet.cell(row=2, column=4).value  # Column D (ADOM)
    
    fortigate_hostnames = []

    # Get FortiGate hostnames from Column A (assuming starting from row 2)
    for row in sheet.iter_rows(min_row=2, max_col=1, max_row=500, values_only=True):  
        Fortigate_hostname = row[0]
        
        if Fortigate_hostname: 
           fortigate_hostnames.append(Fortigate_hostname)

            
    # Close workbook
    workbook.close()

    return fortimanager_ip, api_token, adom, fortigate_hostnames

fortimanager_ip, api_token, adom, fortigate_hostnames = read_excel_data(file_path2)
# API Details
api_url = f"https://{fortimanager_ip}/jsonrpc"
headers = {
    "Authorization": f"Bearer {api_token}",
    "Content-Type": "application/json"
}


def get_fortiswitch_ports(adom, fortigate_hostname):
    resource_path = "/api/v2/cmdb/switch-controller/managed-switch"
    payload = {
        "id": 1,
        "method": "exec",
        "params": [
            {
                "url": "sys/proxy/json",
                "data": {
                    "target": [f"adom/{adom}/device/{fortigate_hostname}"],
                    "action": "get",
                    "resource": resource_path
                }
            }
        ]
    }

    try:
        response = requests.post(api_url, headers=headers, json=payload, verify=False, timeout=100)
        if response.status_code == 200:
            data = response.json()
            
            if "result" in data and data["result"][0]["status"]["code"] == 0:
                switch_info = data["result"][0].get("data", [])
                for switch in switch_info:
                    response_data = switch.get("response", {})
                    if "results" in response_data:
                        results = response_data["results"]
                        for item in results:
                            print(f"Device name : {fortigate_hostname}")
                            print(f"Switch Name: {item.get('name', 'N/A')}")
                            print(f"Serial Number: {item.get('sn', 'N/A')}")
                            print(f"Model: {item.get('model', 'N/A')}")
                            print(f"Version: {item.get('version', 'N/A')}")
                            print("Ports Information:")
                    
                            
                            ports = item.get('ports', [])
                            if ports:
                                for port in ports:
                                    print(f"  Port Name: {port.get('port-name', 'N/A')}")
                                    print(f"    VLAN: {port.get('vlan', 'N/A')}")
                                    print(f"    STATUS: {port.get('status', 'N/A')}")
                                    print(f"    Allowed VLANs: {port.get('allowed-vlans', 'N/A')}")
                                    print(f"    Untagged VLANs: {port.get('untagged-vlans', 'N/A')}")
                                    print(f"    MAC Address: {port.get('mac-addr', 'N/A')}")
                                    print(f"    Export to: {port.get('export-to', 'N/A')}")
                                    print()
                            else:
                                print("No ports found.")
                            print()
            else:
                print(f"Error retrieving switch data: {data}")
        else:
            print(f"HTTP Error {response.status_code} for {resource_path}: {response.text}")

    except Exception as e:
        print(f"Error querying {resource_path}: {e}")


# Run API Query for all FortiGate Hostnames
for fortigate_hostname in fortigate_hostnames:
#Redirect stdout to a file
   file_path = "fortiswitch_output.txt"
   sys.stdout = open(file_path, "a")
   get_fortiswitch_ports(adom, fortigate_hostname)

